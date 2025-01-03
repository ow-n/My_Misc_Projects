"""Alzheimer Data Preprocessing"""
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
# pylint: disable=c0103

class AlzheimerData:
    def __init__(self, save_path):
        self.RES_EXCEL_PATH = save_path

    def CheckIfFileExists(self, file_path):
        """Checks if file exists. If not, exit program."""
        try:
            assert os.path.exists(file_path), f"The file {file_path} does not exist."
        except AssertionError as e:
            print(str(e))
            exit(1)

    def CheckIfFileHasData(self, file_path):
        """Check if file has data. If not, exit program."""
        if os.path.getsize(file_path) == 0:
            print(f"The file {file_path} does not have any data.")
            exit(1)

    def GetMatrix(self, file_path):
        """Loads data with Pandas and converts into Numpy 2d Matrix """
        df = pd.read_excel(file_path, skiprows=1)  # skip row1(labels)
        df = df.apply(pd.to_numeric, errors='coerce')  # convert junk to NaN
        matrix = df.to_numpy()  # convert to 2d Matrix (numpy array)
        print("Data successfully placed into matrix.\n")
        return matrix

    def RemoveColsWithJunk(self, matrix):
        """Remove Cols with more than 70% junk(not real number)"""
        JUNK_THRESHOLD = 0.7
        print("Cleaning Junk Columns")
        print(f"Before | Total cols: {matrix.shape[1]}")
        invalid_cols = np.mean(np.isnan(matrix), axis=0) > JUNK_THRESHOLD
        cleaned_matrix = matrix[:, ~invalid_cols]
        print(f"After  | Total cols: {cleaned_matrix.shape[1]}\n")
        return cleaned_matrix

    def RemoveColsWithZero(self, matrix):
        """Remove Cols with more than 70% zeros"""
        ZERO_THRESHOLD = 0.7
        print("Cleaning Zero Columns")
        print(f"Before | Total cols: {matrix.shape[1]}")
        invalid_cols = np.mean(matrix == 0, axis=0) > ZERO_THRESHOLD
        cleaned_matrix = matrix[:, ~invalid_cols]
        print(f"After  | Total cols: {cleaned_matrix.shape[1]}\n")
        return cleaned_matrix

    def RemoveRowsWithInvalidIC50(self, matrix):
        """
        Remove Rows with invalid IC50 values (NaN, >=20000)
        Note: empty values handled previous with pd.to_numeric(errors='coerce')
        """
        IC50_COL = 0  # IC50 in first column
        IC50_THRESHOLD = 20000
        print("Cleaning IC50 Rows")
        print(f"Before | Total rows: {matrix.shape[0]}")
        ic50_vals = matrix[:, IC50_COL]
        invalid_rows = (pd.isna(ic50_vals) |            # NaN values
                       (ic50_vals >= IC50_THRESHOLD))   # >= 20000
        cleaned_matrix = matrix[~invalid_rows]
        print(f"After  | Total rows: {cleaned_matrix.shape[0]}\n")
        return cleaned_matrix

    def SetOtherJunkToAverageOfColumn (self, matrix):
        """Set all Junk to Avg of its Col & Save to Excel"""
        SHEET_NAME = 'Cleaned'
        print("Setting Junk to Avg of Col")
        junk_count = np.isnan(matrix).sum()
        print(f"Before | Total junks: {junk_count}")
        col_means = np.nanmean(matrix, axis=0)
        matrix = np.where(np.isnan(matrix), col_means, matrix)  # replace NaN to col mean
        junk_count = np.isnan(matrix).sum()
        print(f"After  | Total junks: {junk_count}\n")

        self._save_to_excel(matrix, SHEET_NAME)
        return matrix

    def RescaleData(self, matrix):
        """Rescale matrix cols & Save to Excel"""
        SHEET_NAME = 'Rescaled'
        print("Rescaling matrix...")
        mean_vals = np.mean(matrix, axis=0)
        std_vals = np.std(matrix, axis=0)
        std_vals[std_vals == 0] = 1  # avoid division by 0
        rescaled_matrix = (matrix - mean_vals) / std_vals
        print("\t...Matrix rescaled.\n")

        self._save_to_excel(rescaled_matrix, SHEET_NAME)
        return rescaled_matrix

    def NormalizeData(self, matrix):
        """Normalize matrix cols & Save to Excel"""
        SHEET_NAME = 'Normalized'
        print("Normalizing matrix...")
        min_vals = np.min(matrix, axis=0)
        max_vals = np.max(matrix, axis=0)
        denominators = max_vals - min_vals
        denominators[denominators == 0] = 1  # avoid division by 0
        normalized_matrix = (matrix - min_vals) / denominators
        print("\t...Matrix normalized.\n")

        self._save_to_excel(normalized_matrix, SHEET_NAME)
        return normalized_matrix

    def SortTheNormalizedData(self, matrix):
        """Sort matrix IC50 Ascending & Save to Excel"""
        SHEET_NAME = 'Sorted'
        IC50_COL = 0
        print("Sorting matrix based on IC50 ascending...")
        sorted_matrix = matrix[matrix[:, IC50_COL].argsort()]
        print("\t...matrix sorted.\n")

        self._save_to_excel(sorted_matrix, SHEET_NAME)
        return sorted_matrix

    def GetTrainingData(self, matrix):
        """
        Creates traning matrix & Save to Excel
        Index goes up in steps of 1 then 3 repeating
        eg. Rows: 1, 2, 5, 6, 9, 10
        """
        SHEET_NAME = 'TrainedData'
        print("Creating Training Data...")
        training_indices = [i for i in range(len(matrix)) if (i // 2) % 2 == 0]
        train_matrix = matrix[training_indices]
        print("\t...training Data created.\n")
        self._save_to_excel(train_matrix, SHEET_NAME)
        return train_matrix

    def GetValidationData(self, matrix):
        """
        Creates validation matrix & Save to Excel
        Starts 2nd Row, Goes up in Steps of 4
        eg. Rows: 3, 7, 11, 15
        """
        SHEET_NAME = 'ValidatedData'
        print("Creating Validation Data...")
        validation_matrix = matrix[2::4]
        print("\t...Validation Data created.\n")
        self._save_to_excel(validation_matrix, SHEET_NAME)
        return validation_matrix

    def GetTestingData(self, matrix):
        """
        Use the Sorted matrix to get the Testing data.
        Starts 4th row, Goes up in Steps of 4
        eg. Rows: 4, 8, 12, 16
        """
        SHEET_NAME = 'TestingData'
        print("Creating Testing Data...")
        testing_matrix = matrix[3::4]
        print("\t...Testing Data created.\n")
        self._save_to_excel(testing_matrix, SHEET_NAME)
        return testing_matrix

    def _save_to_excel(self, matrix, sheet_name):
        """Save matrix to Excel as new sheet"""
        if not os.path.exists(self.RES_EXCEL_PATH):
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
        else:
            wb = load_workbook(self.RES_EXCEL_PATH)
        
        ws = wb.create_sheet(sheet_name)
        for row in matrix:
            ws.append(row.tolist())

        wb.save(self.RES_EXCEL_PATH)
        #print(f"Data saved to '{sheet_name}' sheet in {self.RES_EXCEL_PATH}")

def csv_to_xlsx(csv_file, xlsx_file):
    """
    Convert CSV to XLSX
    csv_file: str, path to CSV file
    xlsx_file: str, path to XLSX file
    """
    df = pd.read_csv(csv_file)
    df.to_excel(xlsx_file, index=False)
    exit(0)

def main():
    file_path = 'XandY.xlsx'
    save_path = 'TheResults.xlsx'
    myData = AlzheimerData(save_path)
    myData.CheckIfFileExists(file_path)
    myData.CheckIfFileHasData(file_path)
    originalMatrix = myData.GetMatrix(file_path)
    print(f"Original Matrix Shape: {originalMatrix.shape}\n")
    Matrix1 = myData.RemoveColsWithJunk(originalMatrix)
    Matrix1 = myData.RemoveColsWithZero(Matrix1)
    Matrix1 = myData.RemoveRowsWithInvalidIC50(Matrix1)
    Matrix1 = myData.SetOtherJunkToAverageOfColumn(Matrix1)
    print(f"Cleaned Matrix Shape: {Matrix1.shape}\n")
    Matrix2 = myData.RescaleData(Matrix1)
    Matrix3 = myData.NormalizeData(Matrix2)
    Matrix4 = myData.SortTheNormalizedData(Matrix3)

    Matrix5 = myData.GetTrainingData(Matrix4)
    print(f"Training Matrix Shape: {Matrix5.shape}\n")
    Matrix6 = myData.GetValidationData(Matrix4)
    print(f"Validation Matrix Shape: {Matrix6.shape}\n")
    Matrix7 = myData.GetTestingData(Matrix4)
    print(f"Testing Matrix Shape: {Matrix7.shape}\n")

if __name__ == "__main__":
    main()
